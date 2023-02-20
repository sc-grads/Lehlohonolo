import openpyxl

# Loading excel file
wb = openpyxl.load_workbook('store.xlsx')
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)
sheet = wb['Products']
sheet = wb.active

b2_cell = sheet['B2']
c2_cell = sheet['C2']
print(b2_cell.value, c2_cell.value)
# Another way to access cell with coordinates
print(sheet.cell(row=4, column=2).value)
for a, b, c, d, e in sheet[sheet.dimensions]:
    print(a.value, b.value, c.value, c.value, e.value)
